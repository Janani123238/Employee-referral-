from fastapi import APIRouter, Depends
from fastapi.responses import Response
from sqlalchemy.orm import Session
from sqlalchemy import func
from datetime import datetime, timedelta

from ..database import get_db
from .. import models, schemas
from ..auth import get_current_user, require_hr

router = APIRouter(prefix="/api/reports", tags=["reports"])


def _active_refs():
    return models.Referral.is_deleted.is_(False)


@router.post("/referral-report")
def referral_report(payload: schemas.ReportFilterIn, db: Session = Depends(get_db),
                    user=Depends(require_hr)):
    q = db.query(models.Referral).filter(_active_refs())
    if payload.status:
        q = q.filter(models.Referral.status == payload.status)
    if payload.jobId:
        q = q.filter(models.Referral.job_id == payload.jobId)
    if payload.startDate:
        q = q.filter(models.Referral.submitted_date >= payload.startDate)
    if payload.endDate:
        q = q.filter(models.Referral.submitted_date <= payload.endDate)
    if payload.dept:
        job_ids = [j.id for j in db.query(models.Job).filter(models.Job.dept == payload.dept).all()]
        q = q.filter(models.Referral.job_id.in_(job_ids))
    referrals = q.order_by(models.Referral.submitted_date.desc()).all()

    results = []
    for r in referrals:
        job = db.query(models.Job).filter(models.Job.id == r.job_id).first()
        emp = db.query(models.Employee).filter(models.Employee.id == r.referred_by).first()
        results.append({
            "id": r.id,
            "candidateName": r.candidate_name,
            "email": r.email,
            "phone": r.phone,
            "jobTitle": job.title if job else "—",
            "department": job.dept if job else "—",
            "referredBy": emp.name if emp else "—",
            "status": r.status,
            "aiScore": r.ai_score.get("overall", 0) if r.ai_score else 0,
            "matchPercent": r.match_percent,
            "atsScore": r.ats_score,
            "submittedDate": r.submitted_date,
        })

    summary = {
        "total": len(results),
        "byStatus": {},
    }
    for r in results:
        s = r["status"]
        summary["byStatus"][s] = summary["byStatus"].get(s, 0) + 1

    return {"results": results, "summary": summary}


@router.post("/hiring-analytics")
def hiring_analytics(payload: schemas.ReportFilterIn, db: Session = Depends(get_db),
                     user=Depends(require_hr)):
    q = db.query(models.Referral).filter(_active_refs())
    if payload.startDate:
        q = q.filter(models.Referral.submitted_date >= payload.startDate)
    if payload.endDate:
        q = q.filter(models.Referral.submitted_date <= payload.endDate)
    if payload.dept:
        job_ids = [j.id for j in db.query(models.Job).filter(models.Job.dept == payload.dept).all()]
        q = q.filter(models.Referral.job_id.in_(job_ids))
    all_refs = q.all()

    total_referrals = len(all_refs)
    total_jobs = db.query(models.Job).filter(models.Job.status == "Open", models.Job.is_deleted.is_(False)).count()
    total_employees = db.query(models.Employee).filter(models.Employee.is_active == True).count()
    joined = sum(1 for r in all_refs if r.status == "Joined")
    offers = sum(1 for r in all_refs if r.status in ("Offer", "Offer Released", "Selected"))
    interviewed = sum(1 for r in all_refs if r.status in ("Interview Scheduled", "Interview Completed", "Selected", "Offer", "Offer Released", "Joined"))
    rejected = sum(1 for r in all_refs if r.status == "Rejected")
    inPipeline = total_referrals - joined - rejected

    scores = [r.ai_score.get("overall", 0) for r in all_refs if r.ai_score and r.ai_score.get("overall")]
    avg_score = round(sum(scores) / len(scores)) if scores else 0
    conversionRate = round(joined / total_referrals * 100) if total_referrals else 0
    interviewConversionRate = round(interviewed / total_referrals * 100) if total_referrals else 0
    offerAcceptanceRate = round(joined / offers * 100) if offers else 0
    hiringRate = round(joined / total_referrals * 100) if total_referrals else 0

    # Monthly trends (last 6 months)
    month_map = {}
    now = datetime.now()
    for i in range(5, -1, -1):
        m = (now.replace(day=1) - timedelta(days=i * 31)).replace(day=1)
        key = f"{m.year}-{m.month:02d}"
        label = m.strftime("%b %Y")
        month_map[key] = {"month": label, "referrals": 0, "shortlisted": 0, "interviews": 0, "offers": 0, "joined": 0}
    for r in all_refs:
        if not r.submitted_date:
            continue
        key = f"{r.submitted_date.year}-{r.submitted_date.month:02d}"
        if key not in month_map:
            continue
        month_map[key]["referrals"] += 1
        if r.status in ("Shortlisted", "Interview Scheduled", "Interview Completed", "Selected", "Offer", "Offer Released", "Joined"):
            month_map[key]["shortlisted"] += 1
        if r.status in ("Interview Scheduled", "Interview Completed", "Selected", "Offer", "Offer Released", "Joined"):
            month_map[key]["interviews"] += 1
        if r.status in ("Offer", "Offer Released", "Selected", "Joined"):
            month_map[key]["offers"] += 1
        if r.status == "Joined":
            month_map[key]["joined"] += 1

    # Department performance
    depts = {}
    for r in all_refs:
        job = db.query(models.Job).filter(models.Job.id == r.job_id).first()
        if not job:
            continue
        d = depts.setdefault(job.dept, {"referrals": 0, "joined": 0, "interviews": 0})
        d["referrals"] += 1
        if r.status == "Joined":
            d["joined"] += 1
        if r.status in ("Interview Scheduled", "Interview Completed", "Selected", "Offer", "Offer Released", "Joined"):
            d["interviews"] += 1
    departmentPerformance = []
    for dept, d in depts.items():
        d["conversionRate"] = round(d["joined"] / d["referrals"] * 100) if d["referrals"] else 0
        d["name"] = dept
        departmentPerformance.append(d)

    # Recruiter (referrer) performance
    referrers = {}
    for r in all_refs:
        emp = db.query(models.Employee).filter(models.Employee.id == r.referred_by).first()
        name = emp.name if emp else "—"
        rec = referrers.setdefault(name, {"referrals": 0, "joined": 0, "interviews": 0})
        rec["referrals"] += 1
        if r.status == "Joined":
            rec["joined"] += 1
        if r.status in ("Interview Scheduled", "Interview Completed", "Selected", "Offer", "Offer Released", "Joined"):
            rec["interviews"] += 1
    recruiterPerformance = []
    for name, rec in referrers.items():
        rec["name"] = name
        rec["conversionRate"] = round(rec["joined"] / rec["referrals"] * 100) if rec["referrals"] else 0
        recruiterPerformance.append(rec)
    recruiterPerformance.sort(key=lambda x: x["referrals"], reverse=True)

    topReferrers = {}
    for r in all_refs:
        emp = db.query(models.Employee).filter(models.Employee.id == r.referred_by).first()
        if emp:
            topReferrers[emp.name] = topReferrers.get(emp.name, 0) + 1

    return {
        "totalReferrals": total_referrals,
        "openPositions": total_jobs,
        "activeEmployees": total_employees,
        "joined": joined,
        "offers": offers,
        "interviewed": interviewed,
        "rejected": rejected,
        "inPipeline": inPipeline,
        "avgAiScore": avg_score,
        "conversionRate": conversionRate,
        "interviewConversionRate": interviewConversionRate,
        "offerAcceptanceRate": offerAcceptanceRate,
        "hiringRate": hiringRate,
        "monthlyTrends": list(month_map.values()),
        "departmentPerformance": departmentPerformance,
        "recruiterPerformance": recruiterPerformance,
        "byDepartment": depts,
        "topReferrers": dict(sorted(topReferrers.items(), key=lambda x: x[1], reverse=True)[:10]),
    }


@router.get("/landing-stats")
def landing_stats(db: Session = Depends(get_db)):
    """Public stats for the landing page live counters (no auth)."""
    total_jobs = db.query(models.Job).filter(models.Job.status == "Open", models.Job.is_deleted.is_(False)).count()
    total_referrals = db.query(models.Referral).filter(_active_refs()).count()
    successful_hires = db.query(models.Referral).filter(_active_refs(), models.Referral.status == "Joined").count()

    rewards_paid = 0
    joined_refs = db.query(models.Referral).filter(_active_refs(), models.Referral.status == "Joined").all()
    for r in joined_refs:
        job = db.query(models.Job).filter(models.Job.id == r.job_id).first()
        rewards_paid += job.bonus if job else 0

    return {
        "openJobs": total_jobs,
        "totalReferrals": total_referrals,
        "successfulHires": successful_hires,
        "referralRewardsPaid": rewards_paid,
        "activeEmployees": db.query(models.Employee).filter(models.Employee.is_active == True).count(),
    }


@router.post("/bonus-report")
def bonus_report(payload: schemas.ReportFilterIn, db: Session = Depends(get_db),
                 user=Depends(require_hr)):
    q = db.query(models.Referral).filter(_active_refs(), models.Referral.status == "Joined")
    if payload.startDate:
        q = q.filter(models.Referral.submitted_date >= payload.startDate)
    if payload.endDate:
        q = q.filter(models.Referral.submitted_date <= payload.endDate)
    if payload.dept:
        job_ids = [j.id for j in db.query(models.Job).filter(models.Job.dept == payload.dept).all()]
        q = q.filter(models.Referral.job_id.in_(job_ids))
    joined = q.all()
    results = []
    totalBonus = 0
    for r in joined:
        job = db.query(models.Job).filter(models.Job.id == r.job_id).first()
        emp = db.query(models.Employee).filter(models.Employee.id == r.referred_by).first()
        bonus = job.bonus if job else 0
        totalBonus += bonus
        results.append({
            "candidateName": r.candidate_name,
            "jobTitle": job.title if job else "—",
            "referredBy": emp.name if emp else "—",
            "bonus": bonus,
            "joinedDate": r.submitted_date,
        })
    return {"results": results, "totalBonus": totalBonus, "totalPaid": len(results)}


@router.post("/export")
def export_report(payload: schemas.ReportExportIn, db: Session = Depends(get_db),
                  user=Depends(require_hr)):
    """Server-side report export. Supports 'xlsx' (Excel) and 'pdf'.

    Replaces the old client-side CSV/fake-PDF approach so exports include the
    same filter logic as the on-screen report and produce real files.
    """
    report_type = payload.reportType or "referral"
    fmt = payload.format or "xlsx"
    if fmt not in ("xlsx", "pdf"):
        raise HTTPException(status_code=400, detail="Unsupported format. Use 'xlsx' or 'pdf'.")

    flt = payload.filters
    q = db.query(models.Referral).filter(_active_refs())
    if flt.status:
        q = q.filter(models.Referral.status == flt.status)
    if flt.jobId:
        q = q.filter(models.Referral.job_id == flt.jobId)
    if flt.startDate:
        q = q.filter(models.Referral.submitted_date >= flt.startDate)
    if flt.endDate:
        q = q.filter(models.Referral.submitted_date <= flt.endDate)
    if flt.dept:
        job_ids = [j.id for j in db.query(models.Job).filter(models.Job.dept == flt.dept).all()]
        q = q.filter(models.Referral.job_id.in_(job_ids))

    referrals = q.order_by(models.Referral.submitted_date.desc()).all()

    if report_type == "bonus":
        rows = []
        for r in referrals:
            if r.status != "Joined":
                continue
            job = db.query(models.Job).filter(models.Job.id == r.job_id).first()
            emp = db.query(models.Employee).filter(models.Employee.id == r.referred_by).first()
            rows.append([r.candidate_name, job.title if job else "—", emp.name if emp else "—",
                         job.bonus if job else 0, r.submitted_date.strftime("%Y-%m-%d") if r.submitted_date else ""])
        headers = ["Candidate", "Job", "Referred By", "Bonus (INR)", "Joined"]
        title = "MuraAI Refer — Bonus Report"
    else:
        rows = []
        for r in referrals:
            if report_type == "interview" and r.status not in ("Technical Round", "Manager Round", "HR Round"):
                continue
            if report_type == "offer" and r.status != "Offer":
                continue
            if report_type == "joining" and r.status != "Joined":
                continue
            job = db.query(models.Job).filter(models.Job.id == r.job_id).first()
            emp = db.query(models.Employee).filter(models.Employee.id == r.referred_by).first()
            rows.append([
                r.candidate_name, job.title if job else "—", job.dept if job else "—",
                emp.name if emp else "—", r.status,
                (r.ai_score or {}).get("overall", 0) if r.ai_score else 0,
                r.match_percent or 0,
                r.submitted_date.strftime("%Y-%m-%d") if r.submitted_date else "",
            ])
        headers = ["Candidate", "Job", "Department", "Referred By", "Status", "AI Score", "Match %", "Submitted"]
        title = "MuraAI Refer — Referral Report"

    filename = f"muraai-{report_type}-report-{datetime.now().strftime('%Y%m%d-%H%M%S')}"

    if fmt == "xlsx":
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Report"
        ws.append(headers)
        for row in rows:
            ws.append([str(c) if c is not None else "" for c in row])
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = cell.font.copy(bold=True)
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(14, len(h) + 4)
        import io
        buf = io.BytesIO()
        wb.save(buf)
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}.xlsx"'},
        )

    # PDF
    def _pdf_safe(text):
        if text is None:
            return ""
        return "".join(ch if ord(ch) < 256 else "-" for ch in str(text))

    from fpdf import FPDF
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=14)
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 15)
    pdf.cell(0, 10, _pdf_safe(title), ln=1)
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(0, 6, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  Records: {len(rows)}", ln=1)
    pdf.ln(3)
    col_w = [180 / len(headers) for _ in headers]
    pdf.set_font("Helvetica", "B", 8)
    for i, h in enumerate(headers):
        pdf.cell(col_w[i], 7, h[:24], border=1)
    pdf.ln()
    pdf.set_font("Helvetica", "", 7.5)
    for row in rows:
        if pdf.get_y() > 270:
            pdf.add_page()
            pdf.set_font("Helvetica", "B", 8)
            for i, h in enumerate(headers):
                pdf.cell(col_w[i], 7, h[:24], border=1)
            pdf.ln()
            pdf.set_font("Helvetica", "", 7.5)
        for i, c in enumerate(row):
            pdf.cell(col_w[i], 6, _pdf_safe(c)[:40], border=1)
        pdf.ln()
    return Response(
        content=bytes(pdf.output()),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}.pdf"'},
    )
